import json 
import pandas as pd 
from datetime import datetime 
import os 
import time 
from openpyxl.styles import Font 
from openpyxl import load_workbook
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.utils import get_column_letter, coordinate_to_tuple

project_name = "koong"
project_dir = r"C:\Users\master\Project\koong-nft\assets\translations"
output_dir = r"C:\Users\master\Project\python\translate\output"
def get_filpath(type):    
    now = datetime.now()
    date = str(now.year) + str(now.month).zfill(2) + str(now.day).zfill(2)
    filename = project_name + "_" + type + "_" + date + ".xlsx"
    print(filename)
    return os.path.join(output_dir, filename)



def read_json_file(path):
    with open(path,'r', encoding='utf8') as f:
        data = json.load(f)
    return data 

def check_todo(string):
    if string[-2:] == "  ":
        return True
    return False 

def save_excel(type , to_json):

    if type == "en":
        to_col = "영어"
    elif type == "ja":
        to_col = "일본어"
    elif type == "th":
        to_col = "태국어"

    keys = []
    ko = []
    to = [] 

    for key in to_json: 
        string = to_json[key]
        if check_todo(string):
            keys.append(key)
            ko.append(ko_json[key])
            to.append("")


    excel_data = pd.DataFrame(
        {
            'code': keys,
            '한글': ko,
            to_col :to
        }
    )
    output = get_filpath(type) 
    excel_data.to_excel(output , index=False)
    # 행 넓이 넗히기 

    time.sleep(2)
    wb = load_workbook(output)
    ws = wb.active
    for column_cells in ws.columns:
        new_column_length = max(len(str(cell.value)) for cell in column_cells)
        new_column_letter = (get_column_letter(column_cells[0].column))
        if new_column_length > 0:
            ws.column_dimensions[new_column_letter].width = new_column_length*2

    wb.save(output)





ko_file= project_dir + r"\ko.json"
en_file = project_dir + r"\en.json"
ja_file = project_dir + r"\ja.json"

ko_json = read_json_file(ko_file)
en_json = read_json_file(en_file)
ja_json = read_json_file(ja_file)

ko_keys = [key for key in ko_json]
en_keys = [key for key in en_json]
ja_keys = [key for key in ja_json]

if len(ko_keys) != len(en_keys):
    print("국문과 영문파일길이가 맞지 않습니다.")
    exit
if len(ko_keys) != len(en_keys):
    print("국문과 일본어파일 길이가 맞지 않습니다.")
    exit


save_excel('en',en_json)
save_excel('ja',ja_json)








